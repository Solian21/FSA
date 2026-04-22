from openpyxl import load_workbook, Workbook

# -----------------------------
# STEP 1: READ KEY LOG SHEET
# -----------------------------
keylog_file_path = input("Enter key log path name: ")
workbook = load_workbook(keylog_file_path)
sheet = workbook.active

room_list = []
for row in sheet.iter_rows(values_only=True):
    room_list.append(list(row))

updated_list = []
count = 0

lobbyKeyCode = " "
mailKeyCode = " "
garageKeyCode = " "

for line in room_list:
    if count > 564:
        break

    if line[0] is not None:
        keyLogName = line[0].split(" ")

        if len(keyLogName) > 2 and keyLogName[2] == "Apt":
            lobbyKeyCode = line[1]
            continue

        elif len(keyLogName) > 2 and keyLogName[2] == "Mailbox":
            mailKeyCode = line[1]
            continue

        elif len(keyLogName) > 2 and keyLogName[2] == "Garage":
            garageKeyCode = line[1]
            continue

        else:
            updated_list.append([line[0] + " Apt", line[0], "Apt Key", lobbyKeyCode])
            updated_list.append([line[0] + " Mail", line[0], "Mail Key", mailKeyCode])
            updated_list.append([line[0] + " Garage", line[0], "Garage Key", garageKeyCode])
            updated_list.append([line[0] + " Room", line[0], "Room Key", line[1]])

    count += 1

print("\nCleaned key list:")
for row in updated_list:
    print(row)

# -----------------------------
# STEP 2: READ OCCUPANCY SHEET
# -----------------------------
occupancy_file_path = input("\nEnter Occupancy path Name: ")
workbook = load_workbook(occupancy_file_path)
sheet = workbook.active

occupancy_list = []
for row in sheet.iter_rows(values_only=True):
    occupancy_list.append(list(row))

# -----------------------------
# STEP 3: MATCH RESIDENT NAMES
# -----------------------------
residents_name_updated = []

for i in range(len(occupancy_list)):
    roomName = occupancy_list[i][0]
    name = occupancy_list[i][1]

    if roomName is None:
        continue

    for j in range(len(updated_list)):
        if updated_list[j] is not None:
            if str(roomName) in str(updated_list[j][1]):
                residents_name_updated.append([
                    updated_list[j][0],
                    updated_list[j][1],
                    updated_list[j][2],
                    updated_list[j][3],
                    name
                ])

print("\nFinal matched list:")
for row in residents_name_updated:
    print(row)

# -----------------------------
# STEP 4: SAVE FINAL OUTPUT
# -----------------------------
output_file_path = input("\nEnter the output file path (e.g., output.xlsx): ")

new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Updated List"

headers = [
    "Full Room Space Description",
    "Room Space",
    "Key Type Description",
    "Key Code",
    "Residents Name"
]
new_sheet.append(headers)

for row in residents_name_updated:
    new_sheet.append(row)

new_workbook.save(output_file_path)
print(f"\nData successfully saved to {output_file_path}")