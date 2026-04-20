from openpyxl import load_workbook , Workbook


file_path = input("Enter Occupancy path Name")
workbook = load_workbook(file_path)
sheet = workbook.active


occupancy_list = []
for row in sheet.iter_rows(values_only=True):
    occupancy_list.append(list(row))






file_path = input("Enter Cleaned Sheet path Name")
workbook = load_workbook(file_path)
sheet = workbook.active


final_list = []
for row in sheet.iter_rows(values_only=True):
    final_list.append(list(row))



residents_name_updated = []





for i in range(len(occupancy_list)):
    name = occupancy_list[i][1]
    roomName = occupancy_list[i][0]

    if(roomName == None):
        continue



    for j in range(i+ 1, len(final_list)):
        if final_list[j] != None:
            if(roomName in final_list[j][1]):
                residents_name_updated.append([final_list[j][0] , final_list[j][1]  ,final_list[j][2],final_list[j][3] ,name ])

for s in residents_name_updated:
    print(s)


output_file_path = input("Enter the output file path (e.g., output.xlsx): ")
new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Updated List"
headers = ["Full Room Space Description", "Room Space", "Key Type Description", "Key Code","Residents Name"]
new_sheet.append(headers)
for row in residents_name_updated:
    new_sheet.append(row)
new_workbook.save(output_file_path)
print(f"Data successfully saved to {output_file_path}")






