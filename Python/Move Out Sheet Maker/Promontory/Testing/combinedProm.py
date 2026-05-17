from openpyxl import load_workbook, Workbook


def load_sheet_as_list(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def normalize_text(value):
    """
    Makes text easier to compare.

    Examples:
    132 A Bed 1 -> 132ABED1
    132-A Bed 1 -> 132ABED1
    132.0 A Bed 1 -> 132ABED1
    """
    if value is None:
        return ""

    value = str(value).strip().upper()

    if value.endswith(".0"):
        value = value[:-2]

    value = value.replace(" ", "")
    value = value.replace("-", "")

    return value


def is_blank(value):
    return value is None or str(value).strip() == ""


def remove_building_prefix(value):
    """
    Removes the building prefix.

    Example:
    PWPR-1: 132 A Bed 1 -> 132 A Bed 1
    PWPR-1: 132 Apt -> 132 Apt
    """
    if value is None:
        return ""

    text = str(value).strip()

    if ":" in text:
        text = text.split(":", 1)[1].strip()

    return text


def get_unit_key(value):
    """
    Gets the unit number.

    Examples:
    PWPR-1: 132 A Bed 1 -> 132
    PWPR-1: 132 Apt -> 132
    PWPR-1: 132 Mailbox -> 132
    """
    text = remove_building_prefix(value).upper()
    parts = text.split()

    if len(parts) >= 1:
        return normalize_text(parts[0])

    return ""


def get_bed_key(value):
    """
    Gets the full bed key.

    Examples:
    PWPR-1: 132 A Bed 1 -> 132ABED1
    PWPR-1: 131 D Bed 2 -> 131DBED2
    """
    text = remove_building_prefix(value)
    return normalize_text(text)


def is_apt_key_row(description):
    """
    Detects rows like:
    PWPR-1: 132 Apt
    """
    text = remove_building_prefix(description).upper()
    parts = text.split()

    return len(parts) >= 2 and parts[1] == "APT"


def is_mailbox_key_row(description):
    """
    Detects rows like:
    PWPR-1: 132 Mailbox
    """
    text = remove_building_prefix(description).upper()
    parts = text.split()

    return len(parts) >= 2 and parts[1] == "MAILBOX"


def build_key_lookups(key_log_rows):
    """
    Builds two lookup tables from the key log.

    unit_lookup stores Apt/Mail keys by unit:
        132 -> Apt Key 96, Mail Key 3047

    room_key_lookup stores Room keys by exact bed:
        131ABED1 -> R651
        131ABED2 -> R651
        131DBED1 -> R654
        131DBED2 -> R654
    """
    unit_lookup = {}
    room_key_lookup = {}

    for row in key_log_rows:
        if not row or len(row) < 2:
            continue

        description = row[0]
        key_code = row[1]

        if description is None:
            continue

        unit_key = get_unit_key(description)

        if unit_key == "":
            continue

        if unit_key not in unit_lookup:
            unit_lookup[unit_key] = {
                "apt_key": "",
                "mail_key": ""
            }

        if is_apt_key_row(description):
            unit_lookup[unit_key]["apt_key"] = key_code
            continue

        if is_mailbox_key_row(description):
            unit_lookup[unit_key]["mail_key"] = key_code
            continue

        bed_key = get_bed_key(description)

        if bed_key != "":
            room_key_lookup[bed_key] = key_code

    return unit_lookup, room_key_lookup


def build_final_list_from_occupancy(occupancy_rows, unit_lookup, room_key_lookup):
    """
    Every occupancy row gets added to the final output.

    Each occupancy row creates 3 rows:
    - Apt Key
    - Mail Key
    - Room Key

    If Apt/Mail/Room key information is missing, the output row is still added
    with a blank key code.
    """
    final_list = []
    rooms_needing_key_info = []

    for row in occupancy_rows:
        if not row or len(row) < 2:
            continue

        room_space = row[0]
        resident_name = row[1]

        if room_space is None:
            continue

        room_space = str(room_space).strip()

        # Skip a likely header row if present
        if room_space.strip().upper() in ["ROOM", "ROOM SPACE", "SPACE"]:
            continue

        unit_key = get_unit_key(room_space)
        bed_key = get_bed_key(room_space)

        unit_info = unit_lookup.get(unit_key, {})

        apt_key = unit_info.get("apt_key", "")
        mail_key = unit_info.get("mail_key", "")
        room_key = room_key_lookup.get(bed_key, "")

        missing_parts = []

        if is_blank(apt_key):
            missing_parts.append("Apt Key")

        if is_blank(mail_key):
            missing_parts.append("Mail Key")

        if is_blank(room_key):
            missing_parts.append("Room Key")

        if missing_parts:
            rooms_needing_key_info.append({
                "room_space": room_space,
                "unit_key": unit_key,
                "bed_key": bed_key,
                "resident": resident_name,
                "missing": missing_parts
            })

        final_list.append([
            room_space + " Apt",
            room_space,
            "Apt Key",
            apt_key,
            resident_name
        ])

        final_list.append([
            room_space + " Mail",
            room_space,
            "Mail Key",
            mail_key,
            resident_name
        ])

        final_list.append([
            room_space + " Room",
            room_space,
            "Room Key",
            room_key,
            resident_name
        ])

    return final_list, rooms_needing_key_info


def save_to_workbook(data, output_file_path):
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Updated List"

    headers = [
        "Full Room Space Description",
        "Room Space",
        "Key Type Description",
        "Key Code",
        "Residents Name"
    ]
    new_sheet.append(headers)

    for row in data:
        new_sheet.append(row)

    new_workbook.save(output_file_path)


def print_rooms_needing_key_info(rooms_needing_key_info):
    print("\nRooms needing key information:")
    print("--------------------------------")

    if not rooms_needing_key_info:
        print("None. All occupancy rows had key information.")
        return

    for item in rooms_needing_key_info:
        room_space = item["room_space"]
        unit_key = item["unit_key"]
        bed_key = item["bed_key"]
        resident = item["resident"]
        missing = ", ".join(item["missing"])

        print(
            f"Room Space: {room_space} | "
            f"Unit Key: {unit_key} | "
            f"Bed Key: {bed_key} | "
            f"Resident: {resident} | "
            f"Missing: {missing}"
        )


def main():
    key_log_path = input("Enter Key log path name: ")
    occupancy_path = input("Enter Occupancy path Name: ")
    output_file_path = input("Enter the output file path, for example output.xlsx: ")

    key_log_rows = load_sheet_as_list(key_log_path)
    occupancy_rows = load_sheet_as_list(occupancy_path)

    unit_lookup, room_key_lookup = build_key_lookups(key_log_rows)

    final_list, rooms_needing_key_info = build_final_list_from_occupancy(
        occupancy_rows,
        unit_lookup,
        room_key_lookup
    )

    save_to_workbook(final_list, output_file_path)

    print(f"\nData successfully saved to {output_file_path}")
    print(f"Total rows added, not counting header: {len(final_list)}")
    print(f"Total occupancy rows processed: {len(final_list) // 3}")

    print_rooms_needing_key_info(rooms_needing_key_info)


if __name__ == "__main__":
    main()