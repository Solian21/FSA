from openpyxl import load_workbook

print("This will be the sheet that contains all the correct information given from the README")
# Making this so you can just paste without having to get rid of the quotes on the filepath and can literally paste it in the terminal.
file_path = input("Enter path name: ").strip().strip('"').strip("'").replace("\\", "/")
workbook = load_workbook(file_path)
sheet = workbook.active


room_list = []
for row in sheet.iter_rows(values_only=True):
    room_list.append(list(row))

unmatched = []
matched = True

for idx, line in enumerate(room_list, start=1):
    # Comparing Column 1 to Column 3
    if line[0] != line[2]:
        matched = False
        # this is going to pass in the row that is unmatched
        unmatched.append((idx, line[0], line[2]))
        break;

if unmatched:
    print("------------------------------------")
    print("\nError: There is a mismatch")
    print(f"Unmatched rows: {unmatched[0][0]}")
else:
    print("Success: All the rooms matched!")
