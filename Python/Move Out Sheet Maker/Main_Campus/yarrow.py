from openpyxl import load_workbook , Workbook


file_path = input("Enter key log path name")
workbook = load_workbook(file_path)
sheet = workbook.active


room_list = []
for row in sheet.iter_rows(values_only=True):
    room_list.append(list(row))

unmatched = []
matched = True


updated_list= []
count = 0

mailKeyCode = " "


roomWithNoMailKey = {
    "YARH-1: 109",
    "YARH-1: 110 ",
    "YARH-1: 111",
    "YARH-1: 120",
    "YARH-1: 121",
    "YARH-1: 122",
    "YARH-2: 208",
    "YARH-2: 209",
    "YARH-2: 210",
    "YARH-2: 211",
    "YARH-2: 218",
    "YARH-2: 219",
    "YARH-3: 302",
    "YARH-3: 303",
    "YARH-3: 304",
    "YARH-3: 313",
    "YARH-3: 314",
    "YARH-3: 315",
    "YARH-3: 324",
    "YARH-3: 325",
    "YARH-3: 326"
}










for line in room_list:
    if(count > 227):
        break
    if line[0] is not None:


        keyLogName = line[0].split(" ")

        if len(keyLogName) > 2 and keyLogName[2] == "Mailbox":
            mailKeyCode = line[1]
            continue
        elif "Bed" in  keyLogName[2] and keyLogName[0] +" " +keyLogName[1] not in roomWithNoMailKey:
            updated_list.append([line[0] + " Mail",line[0], "Mail Key",mailKeyCode])
            updated_list.append([line[0] + " Room", line[0], "Room Key", line[1]])
        else :
            updated_list.append([line[0], line[0], "Room Key", line[1]])


    count+=1


for s in updated_list:
   print(s)


output_file_path = input("Enter the output file path (e.g., output.xlsx): ")
new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Updated List"

for row in updated_list:
    new_sheet.append(row)
new_workbook.save(output_file_path)
print(f"Data successfully saved to {output_file_path}")





