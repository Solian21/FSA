from openpyxl import load_workbook, Workbook
import re


def normalize_text(value):
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("\xa0", " ")

    # Example: "224-A Bed 2" -> "224 A Bed 2"
    text = re.sub(r"(\d+)-([A-Za-z])", r"\1 \2", text)

    # Collapse extra spaces
    text = re.sub(r"\s+", " ", text)

    return text


def get_room_name(name):
    """
    Gets the shared room name.

    Examples:
    'YARH-1: 101 Bed 1'      -> 'YARH-1: 101'
    'YARH-1: 101 Mailbox'    -> 'YARH-1: 101'
    """
    text = normalize_text(name)

    match = re.match(r"^(.*?:\s*\d+)", text)
    if match:
        return match.group(1).strip()

    parts = text.split()
    if len(parts) >= 2:
        return f"{parts[0]} {parts[1]}"

    return ""


def is_mailbox_row(name):
    words = normalize_text(name).lower().split()
    return "mailbox" in words or "mail" in words


def is_header_row(room_name, resident_name=""):
    room_name_lower = normalize_text(room_name).lower()
    resident_name_lower = normalize_text(resident_name).lower()

    room_headers = {
        "room",
        "room space",
        "space",
        "bed space",
        "room/bed",
        "room bed",
    }

    resident_headers = {
        "resident",
        "resident name",
        "residents name",
        "name",
    }

    return room_name_lower in room_headers or resident_name_lower in resident_headers


def load_sheet_as_list(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    return data


def read_key_log(key_log_list):
    """
    Reads the key log and stores available keys.

    Key log expected:
    Column A = key/room name
    Column B = key code
    """
    mailbox_keys = {}
    room_keys = {}

    for line in key_log_list:
        if not line or len(line) < 2:
            continue

        key_name = normalize_text(line[0])
        key_code = normalize_text(line[1])

        if not key_name or not key_code:
            continue

        base_room = get_room_name(key_name)

        if is_mailbox_row(key_name):
            mailbox_keys[base_room] = key_code
        else:
            # Direct room/bed key
            room_keys[key_name] = key_code

    return mailbox_keys, room_keys


def build_output_from_occupancy(
    occupancy_list,
    mailbox_keys,
    room_keys,
    rooms_with_no_mail_key
):
    """
    Builds output from occupancy so residents/beds are not skipped.

    Occupancy expected:
    Column A = room/bed space
    Column B = resident name

    For each occupancy row:
    - Adds Mail Key if available and not exempt
    - Adds Room Key if available
    - Skips whichever key does not exist
    """
    final_rows = []
    missing_mail_keys = []
    missing_room_keys = []

    for row in occupancy_list:
        if not row or len(row) < 1:
            continue

        room_space = normalize_text(row[0])
        resident_name = normalize_text(row[1]) if len(row) > 1 else ""

        if not room_space:
            continue

        if is_header_row(room_space, resident_name):
            continue

        base_room = get_room_name(room_space)

        mailbox_key_code = mailbox_keys.get(base_room, "")
        room_key_code = room_keys.get(room_space, "")

        # Add Mail Key only if this room is supposed to have one
        if base_room not in rooms_with_no_mail_key:
            if mailbox_key_code:
                final_rows.append([
                    f"{room_space} Mail",
                    room_space,
                    "Mail Key",
                    mailbox_key_code,
                    resident_name
                ])
            else:
                missing_mail_keys.append(room_space)

        # Add Room Key only if it exists
        if room_key_code:
            final_rows.append([
                f"{room_space} Room",
                room_space,
                "Room Key",
                room_key_code,
                resident_name
            ])
        else:
            missing_room_keys.append(room_space)

    return final_rows, missing_mail_keys, missing_room_keys


def unique_keep_order(items):
    seen = set()
    unique_items = []

    for item in items:
        if item not in seen:
            seen.add(item)
            unique_items.append(item)

    return unique_items


def save_to_excel(output_file_path, rows):
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Updated List"

    headers = [
        "Full Room Space Description",
        "Room Space",
        "Key Type Description",
        "Key Code",
        "Residents Name"
    ]

    new_sheet.append(headers)

    for row in rows:
        new_sheet.append(row)

    new_workbook.save(output_file_path)


def main():
    key_log_path = input("Enter key log path name: ").strip()
    occupancy_path = input("Enter occupancy path name: ").strip()
    output_file_path = input("Enter the output file path, e.g. output.xlsx: ").strip()

    rooms_with_no_mail_key = {
        "YARH-1: 109",
        "YARH-1: 110",
        "YARH-1: 111",
        "YARH-1: 120",
        "YARH-1: 121",
        "YARH-1: 122",
        "YARH-2: 208",
        "YARH-2: 209",
        "YARH-2: 210",
        "YARH-2: 211",
        "YARH-2: 218",
        "YARH-2: 219",
        "YARH-3: 302",
        "YARH-3: 303",
        "YARH-3: 304",
        "YARH-3: 313",
        "YARH-3: 314",
        "YARH-3: 315",
        "YARH-3: 324",
        "YARH-3: 325",
        "YARH-3: 326",
    }

    key_log_list = load_sheet_as_list(key_log_path)
    occupancy_list = load_sheet_as_list(occupancy_path)

    mailbox_keys, room_keys = read_key_log(key_log_list)

    final_rows, missing_mail_keys, missing_room_keys = build_output_from_occupancy(
        occupancy_list,
        mailbox_keys,
        room_keys,
        rooms_with_no_mail_key
    )

    for row in final_rows:
        print(row)

    missing_mail_keys = unique_keep_order(missing_mail_keys)
    missing_room_keys = unique_keep_order(missing_room_keys)

    if missing_mail_keys:
        print("\nRooms/Beds missing a mailbox key:")
        for item in missing_mail_keys:
            print(item)

    if missing_room_keys:
        print("\nRooms/Beds with no room key found:")
        for item in missing_room_keys:
            print(item)

    save_to_excel(output_file_path, final_rows)

    print(f"\nData successfully saved to {output_file_path}")


if __name__ == "__main__":
    main()