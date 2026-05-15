import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def get_room_name(name):
    parts = normalize_text(name).split()
    if len(parts) >= 2:
        return f"{parts[0]} {parts[1]}"
    return ""


def is_mailbox_row(name):
    return "Mailbox" in normalize_text(name).split()


def is_bed_row(name):
    return "Bed" in normalize_text(name).split()


def load_sheet_as_list(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data


def process_files(key_log_path, occupancy_path, output_file_path):
    key_log_list = load_sheet_as_list(key_log_path)
    occupancy_list = load_sheet_as_list(occupancy_path)

    roomWithNoMailKey = {
        "YARH-1: 109",
        "YARH-1: 110",
        "YARH-1: 111",
        "YARH-1: 120",
        "YARH-1: 121",
        "YARH-1: 122",
        "YARH-2: 208",
        "YARH-2: 209",
        "YARH-2: 210",
        "YARH-2: 211",
        "YARH-2: 218",
        "YARH-2: 219",
        "YARH-3: 302",
        "YARH-3: 303",
        "YARH-3: 304",
        "YARH-3: 313",
        "YARH-3: 314",
        "YARH-3: 315",
        "YARH-3: 324",
        "YARH-3: 325",
        "YARH-3: 326",
    }

    mailbox_keys = {}
    for line in key_log_list:
        if not line or len(line) < 2:
            continue

        key_name = normalize_text(line[0])
        key_code = line[1]

        if not key_name:
            continue

        if is_mailbox_row(key_name):
            room_name = get_room_name(key_name)
            if room_name:
                mailbox_keys[room_name] = key_code

    cleaned_rows = []
    missing_mail_keys = []

    for line in key_log_list:
        if not line or len(line) < 2:
            continue

        key_name = normalize_text(line[0])
        key_code = line[1]

        if not key_name:
            continue

        if is_mailbox_row(key_name):
            continue

        if is_bed_row(key_name):
            room_name = get_room_name(key_name)

            if room_name not in roomWithNoMailKey:
                mail_key_code = mailbox_keys.get(room_name, "")
                if mail_key_code == "":
                    missing_mail_keys.append(key_name)

                cleaned_rows.append([
                    f"{key_name} Mail",
                    key_name,
                    "Mail Key",
                    mail_key_code
                ])

            cleaned_rows.append([
                f"{key_name} Room",
                key_name,
                "Room Key",
                key_code
            ])
        else:
            cleaned_rows.append([
                key_name,
                key_name,
                "Room Key",
                key_code
            ])

    occupancy_map = {}
    for row in occupancy_list:
        if not row or len(row) < 2:
            continue

        room_name = normalize_text(row[0])
        resident_name = normalize_text(row[1])

        if room_name:
            occupancy_map[room_name] = resident_name

    residents_name_updated = []
    unmatched_rooms = []

    for row in cleaned_rows:
        full_room_space_description = row[0]
        room_space = normalize_text(row[1])
        key_type_description = row[2]
        key_code = row[3]
        resident_name = occupancy_map.get(room_space, "")

        if resident_name == "":
            unmatched_rooms.append(room_space)

        residents_name_updated.append([
            full_room_space_description,
            room_space,
            key_type_description,
            key_code,
            resident_name
        ])

    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Updated List"

    headers = [
        "Full Room Space Description",
        "Room Space",
        "Key Type Description",
        "Key Code",
        "Residents Name"
    ]
    new_sheet.append(headers)

    for row in residents_name_updated:
        new_sheet.append(row)

    new_workbook.save(output_file_path)

    unique_missing = sorted(set(missing_mail_keys))
    unique_unmatched = sorted(set(unmatched_rooms))

    return unique_missing, unique_unmatched


def browse_key_log():
    path = filedialog.askopenfilename(
        title="Select Key Log File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if path:
        key_log_entry.delete(0, tk.END)
        key_log_entry.insert(0, path)


def browse_occupancy():
    path = filedialog.askopenfilename(
        title="Select Occupancy File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if path:
        occupancy_entry.delete(0, tk.END)
        occupancy_entry.insert(0, path)


def browse_output():
    path = filedialog.asksaveasfilename(
        title="Save Output File As",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if path:
        output_entry.delete(0, tk.END)
        output_entry.insert(0, path)


def run_process():
    key_log_path = key_log_entry.get().strip()
    occupancy_path = occupancy_entry.get().strip()
    output_file_path = output_entry.get().strip()

    if not key_log_path:
        messagebox.showerror("Missing File", "Please select the Key Log file.")
        return

    if not occupancy_path:
        messagebox.showerror("Missing File", "Please select the Occupancy file.")
        return

    if not output_file_path:
        messagebox.showerror("Missing File", "Please choose an Output file location.")
        return

    try:
        missing_mail, unmatched_rooms = process_files(
            key_log_path,
            occupancy_path,
            output_file_path
        )

        summary_message = f"Completed successfully.\n\nOutput saved to:\n{output_file_path}"

        if missing_mail:
            summary_message += f"\n\nBeds missing mailbox keys: {len(missing_mail)}"

        if unmatched_rooms:
            summary_message += f"\nRooms/Beds with no matching resident: {len(unmatched_rooms)}"

        messagebox.showinfo("Success", summary_message)

    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n\n{e}")


root = tk.Tk()
root.title("Key Log Processor")
root.geometry("700x230")
root.resizable(False, False)

title_label = tk.Label(root, text="Key Log and Occupancy Processor", font=("Arial", 14, "bold"))
title_label.pack(pady=10)

frame = tk.Frame(root)
frame.pack(padx=15, pady=10, fill="both", expand=True)

tk.Label(frame, text="Key Log File:").grid(row=0, column=0, sticky="w", pady=5)
key_log_entry = tk.Entry(frame, width=65)
key_log_entry.grid(row=0, column=1, padx=5)
tk.Button(frame, text="Browse", command=browse_key_log, width=10).grid(row=0, column=2)

tk.Label(frame, text="Occupancy File:").grid(row=1, column=0, sticky="w", pady=5)
occupancy_entry = tk.Entry(frame, width=65)
occupancy_entry.grid(row=1, column=1, padx=5)
tk.Button(frame, text="Browse", command=browse_occupancy, width=10).grid(row=1, column=2)

tk.Label(frame, text="Output File:").grid(row=2, column=0, sticky="w", pady=5)
output_entry = tk.Entry(frame, width=65)
output_entry.grid(row=2, column=1, padx=5)
tk.Button(frame, text="Browse", command=browse_output, width=10).grid(row=2, column=2)

complete_button = tk.Button(root, text="Complete", command=run_process, width=20, height=2, bg="lightgreen")
complete_button.pack(pady=15)

root.mainloop()