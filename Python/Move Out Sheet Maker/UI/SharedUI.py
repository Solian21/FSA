import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
import re


# ---------------------------
# Shared helpers
# ---------------------------

def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def load_sheet_as_list(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data


# ---------------------------
# Buildings with No Mail Key logic
# ---------------------------

def check_rooms(file_path):
    room_list = load_sheet_as_list(file_path)

    unmatched = []

    for idx, line in enumerate(room_list, start=1):
        col1 = line[0] if len(line) > 0 else None
        col3 = line[2] if len(line) > 2 else None

        if col1 != col3:
            unmatched.append((idx, col1, col3))
            break

    return unmatched


# ---------------------------
# Yarrow logic
# ---------------------------

def get_room_name(name):
    parts = normalize_text(name).split()
    if len(parts) >= 2:
        return f"{parts[0]} {parts[1]}"
    return ""


def is_mailbox_row(name):
    return "Mailbox" in normalize_text(name).split()


def is_bed_row(name):
    return "Bed" in normalize_text(name).split()


def process_yarrow_files(key_log_path, occupancy_path, output_file_path):
    key_log_list = load_sheet_as_list(key_log_path)
    occupancy_list = load_sheet_as_list(occupancy_path)

    roomWithNoMailKey = {
        "YARH-1: 109",
        "YARH-1: 110",
        "YARH-1: 111",
        "YARH-1: 120",
        "YARH-1: 121",
        "YARH-1: 122",
        "YARH-2: 208",
        "YARH-2: 209",
        "YARH-2: 210",
        "YARH-2: 211",
        "YARH-2: 218",
        "YARH-2: 219",
        "YARH-3: 302",
        "YARH-3: 303",
        "YARH-3: 304",
        "YARH-3: 313",
        "YARH-3: 314",
        "YARH-3: 315",
        "YARH-3: 324",
        "YARH-3: 325",
        "YARH-3: 326",
    }

    mailbox_keys = {}
    for line in key_log_list:
        if not line or len(line) < 2:
            continue

        key_name = normalize_text(line[0])
        key_code = line[1]

        if not key_name:
            continue

        if is_mailbox_row(key_name):
            room_name = get_room_name(key_name)
            if room_name:
                mailbox_keys[room_name] = key_code

    cleaned_rows = []
    missing_mail_keys = []

    for line in key_log_list:
        if not line or len(line) < 2:
            continue

        key_name = normalize_text(line[0])
        key_code = line[1]

        if not key_name:
            continue

        if is_mailbox_row(key_name):
            continue

        if is_bed_row(key_name):
            room_name = get_room_name(key_name)

            if room_name not in roomWithNoMailKey:
                mail_key_code = mailbox_keys.get(room_name, "")
                if mail_key_code == "":
                    missing_mail_keys.append(key_name)

                cleaned_rows.append([
                    f"{key_name} Mail",
                    key_name,
                    "Mail Key",
                    mail_key_code
                ])

            cleaned_rows.append([
                f"{key_name} Room",
                key_name,
                "Room Key",
                key_code
            ])
        else:
            cleaned_rows.append([
                key_name,
                key_name,
                "Room Key",
                key_code
            ])

    occupancy_map = {}
    for row in occupancy_list:
        if not row or len(row) < 2:
            continue

        room_name = normalize_text(row[0])
        resident_name = normalize_text(row[1])

        if room_name:
            occupancy_map[room_name] = resident_name

    residents_name_updated = []
    unmatched_rooms = []

    for row in cleaned_rows:
        full_room_space_description = row[0]
        room_space = normalize_text(row[1])
        key_type_description = row[2]
        key_code = row[3]
        resident_name = occupancy_map.get(room_space, "")

        if resident_name == "":
            unmatched_rooms.append(room_space)

        residents_name_updated.append([
            full_room_space_description,
            room_space,
            key_type_description,
            key_code,
            resident_name
        ])

    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Updated List"

    headers = [
        "Full Room Space Description",
        "Room Space",
        "Key Type Description",
        "Key Code",
        "Residents Name"
    ]
    new_sheet.append(headers)

    for row in residents_name_updated:
        new_sheet.append(row)

    new_workbook.save(output_file_path)

    unique_missing = sorted(set(missing_mail_keys))
    unique_unmatched = sorted(set(unmatched_rooms))

    return unique_missing, unique_unmatched


# ---------------------------
# East Campus logic
# ---------------------------

def extract_unit_bed_letter_and_bed_number(space_text):
    if space_text is None:
        return None, None, None

    text = str(space_text).strip()
    match = re.search(r'(\d+)\s+([A-Z])\s+Bed\s+(\d+)', text, re.IGNORECASE)
    if match:
        return match.group(1), match.group(2).upper(), match.group(3)

    return None, None, None


def parse_east_campus_keylog_row(text):
    if text is None:
        return None, None, None

    text = str(text).strip()

    match = re.match(r'^(\d+)\s+Front\s+Door$', text, re.IGNORECASE)
    if match:
        return match.group(1), "FRONT_DOOR", None

    match = re.match(r'^(\d+)\s+Mailbox$', text, re.IGNORECASE)
    if match:
        return match.group(1), "MAIL", None

    match = re.match(r'^(\d+)\s+Garage$', text, re.IGNORECASE)
    if match:
        return match.group(1), "GARAGE", None

    match = re.match(r'^(\d+)\s+([A-Z])\s+Bedroom$', text, re.IGNORECASE)
    if match:
        return match.group(1), "BEDROOM", match.group(2).upper()

    return None, None, None


def process_east_campus_files(key_log_path, occupancy_path, output_file_path):
    key_log_list = load_sheet_as_list(key_log_path)
    occupancy_list = load_sheet_as_list(occupancy_path)

    unit_keys = {}
    bedroom_keys = {}

    for row in key_log_list:
        key_name = row[0] if len(row) > 0 else None
        key_code = row[1] if len(row) > 1 else None

        if key_name is None:
            continue

        unit, key_type, bed_letter = parse_east_campus_keylog_row(key_name)

        if unit is None:
            continue

        if unit not in unit_keys:
            unit_keys[unit] = {"FRONT_DOOR": "", "MAIL": "", "GARAGE": ""}

        if key_type == "FRONT_DOOR":
            unit_keys[unit]["FRONT_DOOR"] = key_code
        elif key_type == "MAIL":
            unit_keys[unit]["MAIL"] = key_code
        elif key_type == "GARAGE":
            unit_keys[unit]["GARAGE"] = key_code
        elif key_type == "BEDROOM" and bed_letter is not None:
            bedroom_keys[(unit, bed_letter)] = key_code

    residents_name_updated = []

    for row in occupancy_list:
        occupancy_space = row[0] if len(row) > 0 else None
        resident_status = row[1] if len(row) > 1 else None

        if occupancy_space is None:
            continue

        occupancy_space = str(occupancy_space).strip()

        if resident_status is None:
            resident_status = ""
        else:
            resident_status = str(resident_status).strip()

        unit, bed_letter, bed_number = extract_unit_bed_letter_and_bed_number(occupancy_space)

        if unit is None:
            continue

        shared = unit_keys.get(unit, {"FRONT_DOOR": "", "MAIL": "", "GARAGE": ""})

        if bed_letter is not None and bed_number is not None:
            full_space = f"{unit} {bed_letter} Bed {bed_number}"
        else:
            full_space = unit

        residents_name_updated.append([
            f"{full_space} Front Door",
            occupancy_space,
            "Front Door Key",
            shared.get("FRONT_DOOR", ""),
            resident_status
        ])

        residents_name_updated.append([
            f"{full_space} Mail",
            occupancy_space,
            "Mail Key",
            shared.get("MAIL", ""),
            resident_status
        ])

        residents_name_updated.append([
            f"{full_space} Garage",
            occupancy_space,
            "Garage Key",
            shared.get("GARAGE", ""),
            resident_status
        ])

        if bed_letter is not None and bed_number is not None:
            bedroom_key_code = bedroom_keys.get((unit, bed_letter), "")
            residents_name_updated.append([
                f"{full_space} Room",
                occupancy_space,
                "Room Key",
                bedroom_key_code,
                resident_status
            ])

    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Updated List"

    headers = [
        "Full Room Space Description",
        "Room Space",
        "Key Type Description",
        "Key Code",
        "Residents Name"
    ]
    new_sheet.append(headers)

    for row in residents_name_updated:
        new_sheet.append(row)

    new_workbook.save(output_file_path)


# ---------------------------
# UI functions
# ---------------------------

def clear_root():
    for widget in root.winfo_children():
        widget.destroy()


def show_start_page():
    clear_root()
    root.title("Building Selection")
    root.geometry("700x300")
    root.resizable(False, False)

    title_label = tk.Label(root, text="Select Building", font=("Arial", 16, "bold"))
    title_label.pack(pady=25)

    subtitle_label = tk.Label(
        root,
        text="Choose which building process you want to run.",
        font=("Arial", 11)
    )
    subtitle_label.pack(pady=5)

    button_frame = tk.Frame(root)
    button_frame.pack(pady=30)

    no_mail_button = tk.Button(
        button_frame,
        text="Buildings with No Mail Key",
        width=24,
        height=2,
        bg="lightblue",
        command=show_no_mail_key_page
    )
    no_mail_button.grid(row=0, column=0, padx=10, pady=5)

    yarrow_button = tk.Button(
        button_frame,
        text="Yarrow",
        width=24,
        height=2,
        bg="#fff9c4",
        command=show_yarrow_page
    )
    yarrow_button.grid(row=0, column=1, padx=10, pady=5)

    prom_button = tk.Button(
        button_frame,
        text="Prom",
        width=24,
        height=2,
        bg="#e1bee7",
        command=show_prom_page
    )
    prom_button.grid(row=1, column=0, padx=10, pady=5)

    east_campus_button = tk.Button(
        button_frame,
        text="East Campus",
        width=24,
        height=2,
        bg="orange",
        command=show_east_campus_page
    )
    east_campus_button.grid(row=1, column=1, padx=10, pady=5)


# ---------------------------
# Buildings with No Mail Key page
# ---------------------------

def browse_cypress_file():
    path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if path:
        cypress_file_entry.delete(0, tk.END)
        cypress_file_entry.insert(0, path)


def run_cypress_check():
    file_path = cypress_file_entry.get().strip()

    if not file_path:
        messagebox.showerror("Missing File", "Please select an Excel file.")
        return

    try:
        unmatched = check_rooms(file_path)

        if unmatched:
            row_num, col1, col3 = unmatched[0]
            messagebox.showerror(
                "Mismatch Found",
                f"There is a mismatch.\n\n"
                f"Row: {row_num}\n"
                f"Column 1: {col1}\n"
                f"Column 3: {col3}"
            )
        else:
            messagebox.showinfo("Success", "All the rooms matched!")

    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n\n{e}")


def show_no_mail_key_page():
    clear_root()

    root.title("Buildings with No Mail Key - Room Match Checker")
    root.geometry("650x220")
    root.resizable(False, False)

    title_label = tk.Label(
        root,
        text="Buildings with No Mail Key - Room Match Checker",
        font=("Arial", 14, "bold")
    )
    title_label.pack(pady=10)

    frame = tk.Frame(root)
    frame.pack(padx=15, pady=10, fill="both", expand=True)

    tk.Label(frame, text="Excel File:").grid(row=0, column=0, sticky="w", pady=5)

    global cypress_file_entry
    cypress_file_entry = tk.Entry(frame, width=60)
    cypress_file_entry.grid(row=0, column=1, padx=5)

    tk.Button(
        frame,
        text="Browse",
        command=browse_cypress_file,
        width=10
    ).grid(row=0, column=2)

    button_frame = tk.Frame(root)
    button_frame.pack(pady=15)

    tk.Button(
        button_frame,
        text="Back",
        command=show_start_page,
        width=14
    ).grid(row=0, column=0, padx=10)

    tk.Button(
        button_frame,
        text="Complete",
        command=run_cypress_check,
        width=18,
        height=2,
        bg="lightgreen"
    ).grid(row=0, column=1, padx=10)


# ---------------------------
# Yarrow page
# ---------------------------

def browse_yarrow_key_log():
    path = filedialog.askopenfilename(
        title="Select Key Log File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if path:
        yarrow_key_log_entry.delete(0, tk.END)
        yarrow_key_log_entry.insert(0, path)


def browse_yarrow_occupancy():
    path = filedialog.askopenfilename(
        title="Select Occupancy File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if path:
        yarrow_occupancy_entry.delete(0, tk.END)
        yarrow_occupancy_entry.insert(0, path)


def browse_yarrow_output():
    path = filedialog.asksaveasfilename(
        title="Save Output File As",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if path:
        yarrow_output_entry.delete(0, tk.END)
        yarrow_output_entry.insert(0, path)


def run_yarrow_process():
    key_log_path = yarrow_key_log_entry.get().strip()
    occupancy_path = yarrow_occupancy_entry.get().strip()
    output_file_path = yarrow_output_entry.get().strip()

    if not key_log_path:
        messagebox.showerror("Missing File", "Please select the Key Log file.")
        return

    if not occupancy_path:
        messagebox.showerror("Missing File", "Please select the Occupancy file.")
        return

    if not output_file_path:
        messagebox.showerror("Missing File", "Please choose an Output file location.")
        return

    try:
        missing_mail, unmatched_rooms = process_yarrow_files(
            key_log_path,
            occupancy_path,
            output_file_path
        )

        summary_message = f"Completed successfully.\n\nOutput saved to:\n{output_file_path}"

        if missing_mail:
            summary_message += f"\n\nBeds missing mailbox keys: {len(missing_mail)}"

        if unmatched_rooms:
            summary_message += f"\nRooms/Beds with no matching resident: {len(unmatched_rooms)}"

        messagebox.showinfo("Success", summary_message)

    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n\n{e}")


def show_yarrow_page():
    clear_root()
    root.title("Yarrow - Key Log and Occupancy Processor")
    root.geometry("760x260")
    root.resizable(False, False)

    title_label = tk.Label(root, text="Yarrow - Key Log and Occupancy Processor", font=("Arial", 14, "bold"))
    title_label.pack(pady=10)

    frame = tk.Frame(root)
    frame.pack(padx=15, pady=10, fill="both", expand=True)

    tk.Label(frame, text="Key Log File:").grid(row=0, column=0, sticky="w", pady=5)
    global yarrow_key_log_entry
    yarrow_key_log_entry = tk.Entry(frame, width=65)
    yarrow_key_log_entry.grid(row=0, column=1, padx=5)
    tk.Button(frame, text="Browse", command=browse_yarrow_key_log, width=10).grid(row=0, column=2)

    tk.Label(frame, text="Occupancy File:").grid(row=1, column=0, sticky="w", pady=5)
    global yarrow_occupancy_entry
    yarrow_occupancy_entry = tk.Entry(frame, width=65)
    yarrow_occupancy_entry.grid(row=1, column=1, padx=5)
    tk.Button(frame, text="Browse", command=browse_yarrow_occupancy, width=10).grid(row=1, column=2)

    tk.Label(frame, text="Output File:").grid(row=2, column=0, sticky="w", pady=5)
    global yarrow_output_entry
    yarrow_output_entry = tk.Entry(frame, width=65)
    yarrow_output_entry.grid(row=2, column=1, padx=5)
    tk.Button(frame, text="Browse", command=browse_yarrow_output, width=10).grid(row=2, column=2)

    button_frame = tk.Frame(root)
    button_frame.pack(pady=15)

    tk.Button(button_frame, text="Back", command=show_start_page, width=14).grid(row=0, column=0, padx=10)
    tk.Button(button_frame, text="Complete", command=run_yarrow_process, width=18, height=2, bg="lightgreen").grid(row=0, column=1, padx=10)


# ---------------------------
# East Campus page
# ---------------------------

def browse_east_campus_key_log():
    path = filedialog.askopenfilename(
        title="Select Key Log File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if path:
        east_campus_key_log_entry.delete(0, tk.END)
        east_campus_key_log_entry.insert(0, path)


def browse_east_campus_occupancy():
    path = filedialog.askopenfilename(
        title="Select Occupancy File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if path:
        east_campus_occupancy_entry.delete(0, tk.END)
        east_campus_occupancy_entry.insert(0, path)


def browse_east_campus_output():
    path = filedialog.asksaveasfilename(
        title="Save Output File As",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if path:
        east_campus_output_entry.delete(0, tk.END)
        east_campus_output_entry.insert(0, path)


def run_east_campus_process():
    key_log_path = east_campus_key_log_entry.get().strip()
    occupancy_path = east_campus_occupancy_entry.get().strip()
    output_file_path = east_campus_output_entry.get().strip()

    if not key_log_path:
        messagebox.showerror("Missing File", "Please select the Key Log file.")
        return

    if not occupancy_path:
        messagebox.showerror("Missing File", "Please select the Occupancy file.")
        return

    if not output_file_path:
        messagebox.showerror("Missing File", "Please choose an Output file location.")
        return

    try:
        process_east_campus_files(key_log_path, occupancy_path, output_file_path)
        messagebox.showinfo(
            "Success",
            f"Completed successfully.\n\nOutput saved to:\n{output_file_path}"
        )

    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n\n{e}")


def show_east_campus_page():
    clear_root()
    root.title("East Campus - Key Log and Occupancy Processor")
    root.geometry("760x260")
    root.resizable(False, False)

    title_label = tk.Label(root, text="East Campus - Key Log and Occupancy Processor", font=("Arial", 14, "bold"))
    title_label.pack(pady=10)

    frame = tk.Frame(root)
    frame.pack(padx=15, pady=10, fill="both", expand=True)

    tk.Label(frame, text="Key Log File:").grid(row=0, column=0, sticky="w", pady=5)
    global east_campus_key_log_entry
    east_campus_key_log_entry = tk.Entry(frame, width=65)
    east_campus_key_log_entry.grid(row=0, column=1, padx=5)
    tk.Button(frame, text="Browse", command=browse_east_campus_key_log, width=10).grid(row=0, column=2)

    tk.Label(frame, text="Occupancy File:").grid(row=1, column=0, sticky="w", pady=5)
    global east_campus_occupancy_entry
    east_campus_occupancy_entry = tk.Entry(frame, width=65)
    east_campus_occupancy_entry.grid(row=1, column=1, padx=5)
    tk.Button(frame, text="Browse", command=browse_east_campus_occupancy, width=10).grid(row=1, column=2)

    tk.Label(frame, text="Output File:").grid(row=2, column=0, sticky="w", pady=5)
    global east_campus_output_entry
    east_campus_output_entry = tk.Entry(frame, width=65)
    east_campus_output_entry.grid(row=2, column=1, padx=5)
    tk.Button(frame, text="Browse", command=browse_east_campus_output, width=10).grid(row=2, column=2)

    button_frame = tk.Frame(root)
    button_frame.pack(pady=15)

    tk.Button(button_frame, text="Back", command=show_start_page, width=14).grid(row=0, column=0, padx=10)
    tk.Button(button_frame, text="Complete", command=run_east_campus_process, width=18, height=2, bg="lightgreen").grid(row=0, column=1, padx=10)


# ---------------------------
# Placeholder pages
# ---------------------------

def show_prom_page():
    messagebox.showinfo("Coming Soon", "Prom tool is not implemented yet.")


# ---------------------------
# App start
# ---------------------------

root = tk.Tk()
show_start_page()
root.mainloop()