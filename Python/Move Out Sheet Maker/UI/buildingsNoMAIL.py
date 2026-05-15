import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook


def load_sheet_as_list(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    room_list = []
    for row in sheet.iter_rows(values_only=True):
        room_list.append(list(row))

    return room_list


def check_rooms(file_path):
    room_list = load_sheet_as_list(file_path)

    unmatched = []
    matched = True

    for idx, line in enumerate(room_list, start=1):
        col1 = line[0] if len(line) > 0 else None
        col3 = line[2] if len(line) > 2 else None

        if col1 != col3:
            matched = False
            unmatched.append((idx, col1, col3))
            break

    return matched, unmatched


def browse_file():
    path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if path:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, path)


def run_check():
    file_path = file_entry.get().strip()

    if not file_path:
        messagebox.showerror("Missing File", "Please select an Excel file.")
        return

    try:
        matched, unmatched = check_rooms(file_path)

        if unmatched:
            row_num, col1, col3 = unmatched[0]
            messagebox.showerror(
                "Mismatch Found",
                f"There is a mismatch.\n\n"
                f"Row: {row_num}\n"
                f"Column 1: {col1}\n"
                f"Column 3: {col3}"
            )
        else:
            messagebox.showinfo("Success", "All the rooms matched!")

    except Exception as e:
        messagebox.showerror("Error", f"Something went wrong:\n\n{e}")


root = tk.Tk()
root.title("Room Match Checker")
root.geometry("650x180")
root.resizable(False, False)

title_label = tk.Label(root, text="Room Match Checker", font=("Arial", 14, "bold"))
title_label.pack(pady=10)

frame = tk.Frame(root)
frame.pack(padx=15, pady=10, fill="both", expand=True)

tk.Label(frame, text="Excel File:").grid(row=0, column=0, sticky="w", pady=5)
file_entry = tk.Entry(frame, width=60)
file_entry.grid(row=0, column=1, padx=5)
tk.Button(frame, text="Browse", command=browse_file, width=10).grid(row=0, column=2)

complete_button = tk.Button(root, text="Check Matches", command=run_check, width=20, height=2, bg="lightgreen")
complete_button.pack(pady=15)

root.mainloop()