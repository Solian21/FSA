from openpyxl import load_workbook, Workbook
import re


def extract_unit_bed_letter_and_bed_number(space_text):
    if space_text is None:
        return None, None, None

    text = str(space_text).strip()
    match = re.search(r'(\d+)\s+([A-Z])\s+Bed\s+(\d+)', text, re.IGNORECASE)
    if match:
        return match.group(1), match.group(2).upper(), match.group(3)

    return None, None, None


def parse_keylog_row(text):
    if text is None:
        return None, None, None

    text = str(text).strip()

    match = re.match(r'^(\d+)\s+Front\s+Door$', text, re.IGNORECASE)
    if match:
        return match.group(1), "FRONT_DOOR", None

    match = re.match(r'^(\d+)\s+Mailbox$', text, re.IGNORECASE)
    if match:
        return match.group(1), "MAIL", None

    match = re.match(r'^(\d+)\s+Garage$', text, re.IGNORECASE)
    if match:
        return match.group(1), "GARAGE", None

    match = re.match(r'^(\d+)\s+([A-Z])\s+Bedroom$', text, re.IGNORECASE)
    if match:
        return match.group(1), "BEDROOM", match.group(2).upper()

    return None, None, None


# -----------------------------
# STEP 1: READ KEY LOG SHEET
# -----------------------------
keylog_file_path = input("Enter key log path name: ")
workbook = load_workbook(keylog_file_path)
sheet = workbook.active

unit_keys = {}
bedroom_keys = {}

for row in sheet.iter_rows(values_only=True):
    key_name = row[0] if len(row) > 0 else None
    key_code = row[1] if len(row) > 1 else None

    if key_name is None:
        continue

    unit, key_type, bed_letter = parse_keylog_row(key_name)

    if unit is None:
        continue

    if unit not in unit_keys:
        unit_keys[unit] = {"FRONT_DOOR": "", "MAIL": "", "GARAGE": ""}

    if key_type == "FRONT_DOOR":
        unit_keys[unit]["FRONT_DOOR"] = key_code
    elif key_type == "MAIL":
        unit_keys[unit]["MAIL"] = key_code
    elif key_type == "GARAGE":
        unit_keys[unit]["GARAGE"] = key_code
    elif key_type == "BEDROOM" and bed_letter is not None:
        bedroom_keys[(unit, bed_letter)] = key_code


# -----------------------------
# STEP 2: READ OCCUPANCY SHEET
# -----------------------------
occupancy_file_path = input("Enter Occupancy path Name: ")
workbook = load_workbook(occupancy_file_path)
sheet = workbook.active

residents_name_updated = []

for row in sheet.iter_rows(values_only=True):
    occupancy_space = row[0] if len(row) > 0 else None
    resident_status = row[1] if len(row) > 1 else None

    if occupancy_space is None:
        continue

    occupancy_space = str(occupancy_space).strip()

    if resident_status is None:
        resident_status = ""
    else:
        resident_status = str(resident_status).strip()

    unit, bed_letter, bed_number = extract_unit_bed_letter_and_bed_number(occupancy_space)

    if unit is None:
        continue

    shared = unit_keys.get(unit, {"FRONT_DOOR": "", "MAIL": "", "GARAGE": ""})

    if bed_letter is not None and bed_number is not None:
        full_space = f"{unit} {bed_letter} Bed {bed_number}"
    else:
        full_space = unit

    residents_name_updated.append([
        f"{full_space} Front Door",
        occupancy_space,
        "Front Door Key",
        shared.get("FRONT_DOOR", ""),
        resident_status
    ])

    residents_name_updated.append([
        f"{full_space} Mail",
        occupancy_space,
        "Mail Key",
        shared.get("MAIL", ""),
        resident_status
    ])

    residents_name_updated.append([
        f"{full_space} Garage",
        occupancy_space,
        "Garage Key",
        shared.get("GARAGE", ""),
        resident_status
    ])

    if bed_letter is not None and bed_number is not None:
        bedroom_key_code = bedroom_keys.get((unit, bed_letter), "")
        residents_name_updated.append([
            f"{full_space} Room",
            occupancy_space,
            "Room Key",
            bedroom_key_code,
            resident_status
        ])

# -----------------------------
# STEP 3: SAVE FINAL OUTPUT
# -----------------------------
output_file_path = input("\nEnter the output file path (e.g., output.xlsx): ")

new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Updated List"

headers = [
    "Full Room Space Description",
    "Room Space",
    "Key Type Description",
    "Key Code",
    "Residents Name"
]
new_sheet.append(headers)

for row in residents_name_updated:
    new_sheet.append(row)

new_workbook.save(output_file_path)
print(f"\nData successfully saved to {output_file_path}")