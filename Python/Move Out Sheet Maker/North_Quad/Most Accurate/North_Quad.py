from openpyxl import load_workbook, Workbook
import re


def read_sheet_to_list(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def normalize_room_name(value):
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("\xa0", " ")

    # Converts "224-A" to "224 A"
    text = re.sub(r"(\d+)-([A-Z])", r"\1 \2", text)

    # Collapse extra spaces
    text = re.sub(r"\s+", " ", text)

    return text


def get_base_room(room_name):
    """
    Examples:
    VISTE-1: 107 Lobby      -> VISTE-1: 107
    VISTE-1: 107 A Bed 1    -> VISTE-1: 107
    VISTE-1: 109 D Bed 2    -> VISTE-1: 109
    """
    room_name = normalize_room_name(room_name)

    match = re.match(r"^(.*?:\s*\d+)", room_name)
    if match:
        return match.group(1).strip()

    return room_name


def get_room_group(room_name):
    """
    Keeps the letter when a room has A/B/C/D spaces.

    Examples:
    VISTE-1: 107 A Bed 1    -> VISTE-1: 107 A
    VISTE-1: 107 C Mail     -> VISTE-1: 107 C
    VISTE-1: 107 Mail       -> VISTE-1: 107
    VISTE-1: 107 Lobby      -> VISTE-1: 107
    """
    room_name = normalize_room_name(room_name)

    match = re.match(r"^(.*?:\s*\d+)(?:\s+([A-Z]))?", room_name)
    if match:
        base_room = match.group(1).strip()
        letter = match.group(2)

        if letter:
            return f"{base_room} {letter}"

        return base_room

    return room_name


def is_lobby_row(room_name):
    room_name = normalize_room_name(room_name).lower()
    return room_name.endswith(" lobby")


def is_mail_row(room_name):
    room_name = normalize_room_name(room_name).lower()
    return room_name.endswith(" mail") or room_name.endswith(" mailbox")


def read_key_log(room_list):
    """
    Reads the key log into lookup dictionaries.

    Key log expected:
    Column A = room/key name
    Column B = key code
    """
    lobby_keys = {}
    mail_keys = {}
    room_keys = {}

    for line in room_list:
        if not line or len(line) < 2 or line[0] is None:
            continue

        room_name = normalize_room_name(line[0])
        key_code = line[1]

        if not room_name or key_code in (None, ""):
            continue

        base_room = get_base_room(room_name)
        room_group = get_room_group(room_name)

        if is_lobby_row(room_name):
            # Lobby keys still belong to the base room.
            lobby_keys[base_room] = key_code

        elif is_mail_row(room_name):
            # Mail keys can belong to either:
            # VISTE-1: 107
            # or VISTE-1: 107 A / B / C / D
            mail_keys[room_group] = key_code

        else:
            room_keys[room_name] = key_code

    return lobby_keys, mail_keys, room_keys


def build_rows_from_occupancy(occupancy_list, lobby_keys, mail_keys, room_keys):
    """
    Builds output from occupancy list so nobody is skipped.

    Occupancy file expected:
    Column A = room/bed space
    Column B = resident name
    """
    final_rows = []

    for i in range(len(occupancy_list)):
        occ_row = occupancy_list[i]

        if not occ_row or len(occ_row) < 1:
            continue

        room_space = normalize_room_name(occ_row[0])
        resident_name = occ_row[1] if len(occ_row) > 1 else ""

        if not room_space:
            continue

        base_room = get_base_room(room_space)
        room_group = get_room_group(room_space)

        lobby_key_code = lobby_keys.get(base_room)

        # Try the A/B/C/D mail key first.
        # If that does not exist, fall back to the base room mail key.
        mail_key_code = mail_keys.get(room_group) or mail_keys.get(base_room)

        room_key_code = room_keys.get(room_space)

        if lobby_key_code:
            final_rows.append([
                f"{room_space} Lobby",
                room_space,
                "Lobby Key",
                lobby_key_code,
                resident_name
            ])

        if mail_key_code:
            final_rows.append([
                f"{room_space} Mail",
                room_space,
                "Mail Key",
                mail_key_code,
                resident_name
            ])

        if room_key_code:
            final_rows.append([
                f"{room_space} Room",
                room_space,
                "Room Key",
                room_key_code,
                resident_name
            ])

    return final_rows


def save_to_excel(output_file_path, rows):
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Updated List"

    headers = [
        "Full Room Space Description",
        "Room Space",
        "Key Type Description",
        "Key Code",
        "Residents Name"
    ]

    new_sheet.append(headers)

    for row in rows:
        new_sheet.append(row)

    new_workbook.save(output_file_path)


def main():
    key_log_path = input("Enter key log path name: ").strip()
    occupancy_path = input("Enter occupancy path name: ").strip()
    output_file_path = input("Enter the output file path, e.g. output.xlsx: ").strip()

    key_log_rows = read_sheet_to_list(key_log_path)
    occupancy_rows = read_sheet_to_list(occupancy_path)

    lobby_keys, mail_keys, room_keys = read_key_log(key_log_rows)

    final_rows = build_rows_from_occupancy(
        occupancy_rows,
        lobby_keys,
        mail_keys,
        room_keys
    )

    # for row in final_rows:
        # print(row)

    save_to_excel(output_file_path, final_rows)

    print(f"Data successfully saved to {output_file_path}")


if __name__ == "__main__":
    main()