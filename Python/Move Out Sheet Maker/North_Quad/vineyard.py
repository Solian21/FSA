from openpyxl import load_workbook , Workbook


file_path = input("Enter key log path name")
workbook = load_workbook(file_path)
sheet = workbook.active


room_list = []
for row in sheet.iter_rows(values_only=True):
    room_list.append(list(row))

unmatched = []
matched = True


updated_list= []
count = 0

lobbyKeyCode = " "

for line in room_list:
    if(count > 564):
        break
    if line[0] is not None:


        keyLogName = line[0].split(" ")
        #print(keyLogName)
        if len(keyLogName) > 2 and keyLogName[2] == "Lobby":
            lobbyKeyCode = line[1]
            continue
        else :
            updated_list.append([line[0] + " Suite",line[0], "Suite Key",lobbyKeyCode])
            updated_list.append([line[0] + " Room", line[0], "Room Key", line[1]])



    count+=1

for s in updated_list:
    print(s)



output_file_path = input("Enter the output file path (e.g., output.xlsx): ")
new_workbook = Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Updated List"
headers = ["Full Room Space Description", "Room Space", "Key Type Description", "Key Code","Residents Name"]
new_sheet.append(headers)
for row in updated_list:
    new_sheet.append(row)
new_workbook.save(output_file_path)
print(f"Data successfully saved to {output_file_path}")





