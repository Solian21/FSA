from openpyxl import load_workbook, Workbook


def read_sheet_to_list(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def build_cleaned_key_list(room_list, max_rows=565):
    updated_list = []
    lobby_key_code = ""
    mail_key_code = ""

    count = 0

    for line in room_list:
        if count >= max_rows:
            break

        if not line or line[0] is None:
            count += 1
            continue

        room_name = str(line[0]).strip()
        room_key_code = line[1] if len(line) > 1 else None
        key_log_name = room_name.split()

        if len(key_log_name) > 2 and key_log_name[2] == "Lobby":
            lobby_key_code = room_key_code
        elif len(key_log_name) > 3 and key_log_name[3] == "Mailbox":
            mail_key_code = room_key_code
        else:
            updated_list.append([f"{room_name} Lobby", room_name, "Lobby Key", lobby_key_code])
            updated_list.append([f"{room_name} Mail", room_name, "Mail Key", mail_key_code])
            updated_list.append([f"{room_name} Room", room_name, "Room Key", room_key_code])

        count += 1

    return updated_list


def match_residents(occupancy_list, cleaned_list):
    residents_name_updated = []

    # assume row 0 is headers in occupancy and cleaned sheet
    for i in range(1, len(occupancy_list)):
        occ_row = occupancy_list[i]

        if not occ_row or len(occ_row) < 2:
            continue

        room_name = occ_row[0]
        resident_name = occ_row[1]

        if room_name is None:
            continue

        room_name = str(room_name).strip()

        for j in range(len(cleaned_list)):
            clean_row = cleaned_list[j]

            if not clean_row or len(clean_row) < 4:
                continue

            clean_room_space = clean_row[1]

            if clean_room_space is not None and room_name == str(clean_room_space).strip():
                residents_name_updated.append([
                    clean_row[0],   # Full Room Space Description
                    clean_row[1],   # Room Space
                    clean_row[2],   # Key Type Description
                    clean_row[3],   # Key Code
                    resident_name   # Residents Name
                ])

    return residents_name_updated


def save_to_excel(output_file_path, rows):
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Updated List"

    headers = [
        "Full Room Space Description",
        "Room Space",
        "Key Type Description",
        "Key Code",
        "Residents Name"
    ]
    new_sheet.append(headers)

    for row in rows:
        new_sheet.append(row)

    new_workbook.save(output_file_path)


def main():
    key_log_path = input("Enter key log path name: ").strip()
    occupancy_path = input("Enter occupancy path name: ").strip()
    output_file_path = input("Enter the output file path (e.g., output.xlsx): ").strip()

    # Step 1: Read key log and build cleaned list
    room_list = read_sheet_to_list(key_log_path)
    cleaned_list = build_cleaned_key_list(room_list)

    # Step 2: Read occupancy and match residents
    occupancy_list = read_sheet_to_list(occupancy_path)
    final_rows = match_residents(occupancy_list, cleaned_list)

    # Print results
    for row in final_rows:
        print(row)

    # Step 3: Save final output
    save_to_excel(output_file_path, final_rows)
    print(f"Data successfully saved to {output_file_path}")


if __name__ == "__main__":
    main()