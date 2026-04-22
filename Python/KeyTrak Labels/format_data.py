import openpyxl as xl

# Load original workbook
work_book = xl.load_workbook('./Data/FP1_FP2_(unformated).xlsx')
active_sheet = work_book.active

print(active_sheet['A1'].value)

# Get Column A values
column_A = active_sheet['A']

data = []
for cell in column_A:
    if cell.value is not None:
        data.append(cell.value)

print('First 4 of data:', data[:4])

# Split data
split_data = []
for cell in data:
    split_data.append(cell.split())

print('\nFirst 4 of split data:', split_data[:4])

# Clean and keep unique based ONLY on Column B
unique_B_values = set()
cleaned_data = []

for row in split_data:
    if len(row) < 2:
        continue

    colA = row[0]
    colB = row[1]

    # Clean B value (example: 3402-B -> 3402)
    colB_clean = colB.split("-")[0].strip()

    # Keep only unique Column B values
    if colB_clean not in unique_B_values:
        unique_B_values.add(colB_clean)
        cleaned_data.append([colA, colB_clean])

print('\nFirst 4 of cleaned data:', cleaned_data[:4])
print('Total unique values:', len(cleaned_data))

# -------------------------
# Create NEW Excel Workbook
# -------------------------

new_workbook = xl.Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Updated_EC_Data"

# Optional headers
new_sheet["A1"] = "Code"
new_sheet["B1"] = "Value"

# Write cleaned unique data
for row_index, row_data in enumerate(cleaned_data, start=2):
    new_sheet.cell(row=row_index, column=1).value = row_data[0]
    new_sheet.cell(row=row_index, column=2).value = row_data[1]

# Save new file (does NOT touch original)
new_workbook.save('./Data/Updated_EC_Data.xlsx')

print("\nNew Excel file created successfully.")
